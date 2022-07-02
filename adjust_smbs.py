from nightscout_ml_base import NightscoutMlBase
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook


class AdjustSmbs(NightscoutMlBase):
    no_insulin_threshold = 70
    no_insulin_when_dropping_threshold = 100
    base_isf = 100
    max_smb = 2

    def adjust_smbs(self, start_date_time_str):
        # df = pd.read_csv('adjustment_test.csv')
        df = pd.read_csv('data/aiSMB_records.csv')

        df['smbToGive'] = df['smbGiven']

        # when below 70 then don't give insulin
        df.loc[df['bg'] < self.no_insulin_threshold, 'smbToGive'] = 0

        # when below 100 & dropping, don't give insulin
        df.loc[(df['bg'] < self.no_insulin_when_dropping_threshold) & (df['delta'] < 0), 'smbToGive'] = 0

        start_date = self.str_to_time(start_date_time_str)
        for index, row in df.iterrows():
            row_date = self.str_to_time(row['dateStr'])
            if row_date > start_date:
                self.adjust_for_lows(index, row, df)
                self.adjust_for_highs(index, row, df)
        
        df['adjustment'] = df['smbToGive'] - df['smbGiven']

        start_index = df.index[df['dateStr'] == start_date_time_str].tolist()[0]
        df = df[start_index:]

        # df.to_csv('data/adjustment_test_updated.csv', index=False)
        # df.to_csv('data/aiSMB_records_adjusted.code.csv', index=False)
        book = load_workbook('data.xlsx')
        writer = pd.ExcelWriter('data.xlsx', engine='openpyxl')
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}
        start_row = writer.sheets['training_data'].max_row - 1
        # data = pd.read_excel('data.xlsx','training_data')
        # start_row = data.index[df['dateStr'] == start_date_time_str].tolist()[0]
        # 6/29/22 03:10PM - 9850
        df.to_excel(writer, sheet_name='training_data', startrow=start_row, index = False, header= False)

        writer.save()

    def adjust_for_lows(self, index, row, df):
        # if low, calculate insulin suplus (delta/isf) - go back 30+ min & delete
        if row['bg'] < (row['targetBg']-20) and row['delta'] < 0:
            u_to_remove = self.u_to_adjust_based_on_delta(row)
            self.remove_prior_insulin(index, row, df, u_to_remove)
        
        # if low treatment - go back 30+ min & delete 1u
        if row['low_treatment'] == 1:
            u_to_remove = .5
            self.remove_prior_insulin(index, row, df, u_to_remove)

    def adjust_for_highs(self, index, row, df):
        trending_up_or_stable = row['delta'] > -2 or row['shortAvgDelta'] > -2
        if row['bg'] > (row['targetBg']+20) and trending_up_or_stable:
            u_to_add = self.u_to_adjust_based_on_delta(row)
            self.add_prior_insulin(index, row, df, u_to_add)

        if row['more_aggressive'] == 1:
            u_to_add = .5
            self.add_prior_insulin(index, row, df, u_to_add)
        
    def add_prior_insulin(self, index, row, df, u_to_add):
        one_hour_ago = index - 12 if index > 12 else 0
        df_last_hour = df[one_hour_ago:index]
        orig_date = self.str_to_time(row['dateStr'])
        min_date = orig_date - pd.DateOffset(minutes=13*5)
        for i, recent_row in df_last_hour.iterrows():
            # don't add insulin if carbs are added later, add after carbs are added
            more_carbs_added_later = recent_row['cob'] == 0 and self.more_carbs_added_later(recent_row['cob'], i, df_last_hour, index)
            # don't add insulin if pending low
            upcoming_low = self.check_for_upcoming_low(i, df_last_hour)
            # don't add if old date
            out_dated = self.str_to_time(recent_row['dateStr']) < min_date

            if not more_carbs_added_later and not upcoming_low and not out_dated:
                current_smb = recent_row['smbToGive']
                if u_to_add + current_smb < self.max_smb:
                    df.at[i, 'smbToGive'] = current_smb + u_to_add
                    return
                if u_to_add < self.max_smb:
                    df.at[i, 'smbToGive'] = 2
                    u_to_add -= self.max_smb - current_smb

    def more_carbs_added_later(self, current_cob, recent_index, df_last_hour, index):
        for i in range(recent_index, index):
            if current_cob < df_last_hour.at[i, 'cob']:
                return True
        return False

    def check_for_upcoming_low(self, recent_index, df_last_hour):
        for i in range(recent_index, len(df_last_hour)):
            row = df_last_hour.iloc[i]
            low_and_dropping = row['bg'] < self.no_insulin_when_dropping_threshold and row['delta'] < 0
            below_threshold = row['bg'] < self.no_insulin_threshold
            if below_threshold or low_and_dropping:
                return True
        return False

    def remove_prior_insulin(self, index, row, df, u_to_remove):
        orig_date = self.str_to_time(row['dateStr'])
        min_date = orig_date - pd.DateOffset(minutes=37*5)

        for i in range(6, 36):
            if index - i > 0:
                prior_row = df.iloc[index - i]
                row_date = self.str_to_time(prior_row['dateStr'])
                if row_date > min_date:
                    smb_given = prior_row['smbToGive']
                    if smb_given >= u_to_remove:
                        new_smb = smb_given - u_to_remove
                        df.at[index-i, 'smbToGive'] = new_smb
                        return
                    if smb_given > 0 and smb_given < u_to_remove:
                        df.at[index-i, 'smbToGive'] = 0
                        u_to_remove -= smb_given

    def u_to_adjust_based_on_delta(self, row):
        units = abs(row['delta']) / self.get_isf(row['bg'])
        units = self.roundToPt05(units)

        if row['bg'] > 150 and 3 > row['delta'] > -3 and 2 > row['longAvgDelta'] > -2 and units == .05:
            units = .25

        # still want to remove some insulin if .02 should be removed
        units = .05 if units == 0 else units
        return units

    def roundToPt05(self, units):
        return round(units * 20) / 20

    def get_isf(self, bg):
        if bg > 150:
            return self.base_isf * .8
        if bg > 200:
            return self.base_isf * .6
        return self.base_isf