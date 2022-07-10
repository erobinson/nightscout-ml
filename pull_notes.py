import sys
sys.path.append('../nightscout-python-client')
import nightscout_python_client
import csv
from datetime import datetime, timedelta
from nightscout_ml_base import NightscoutMlBase
import pandas as pd
from openpyxl import load_workbook

class PullNotes(NightscoutMlBase):
    configuration = nightscout_python_client.Configuration()
    configuration.host = "erjr-nightscout.herokuapp.com/api/v1"
    treatment_api = nightscout_python_client.TreatmentsApi(nightscout_python_client.ApiClient(configuration))
    output_file_path = 'data/recent_notes.csv'
    timezone_offset_from_zulu = 4

    def pull_notes_to_csv(self, start_date_time):
        count = 10000
        now = datetime.now()
        now = now.replace(hour=0, minute=0, second=0, microsecond=0)
        end_date = now + timedelta(days=1)
        start_date = self.str_to_time(start_date_time)
        start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
        
        api_response = self.treatment_api.treatments_get("note", createdAtGte=start_date.isoformat(), createdAtLte=end_date.isoformat(), count=count)
        
        sgv_output_file = open(self.output_file_path, 'w')
        csv_writer = csv.writer(sgv_output_file)
        csv_writer.writerow(['dateStr', 'note'])
        for line in api_response:
            if 'notes' in line and ('low' in line['notes'].lower() or 'aggressive' in line['notes'].lower()):
                created_at_dt = datetime.strptime(line['created_at'], '%Y-%m-%dT%H:%M:%S.%fZ')
                created_at_dt = created_at_dt - timedelta(hours=self.timezone_offset_from_zulu)
                created_at_str = self.time_to_str(created_at_dt)
                csv_writer.writerow([created_at_str,line['notes']])
        
        return self.output_file_path

    def add_adjustment_flags(self, aismb_records_file_str, notes_file_str):
        notes_pd = pd.read_csv(notes_file_str)
        aismb_df = pd.read_csv(aismb_records_file_str)
        
        if 'low_treatment' not in aismb_df:
            aismb_df['low_treatment'] = 0
        
        if 'more_aggressive' not in aismb_df:
            aismb_df['more_aggressive'] = 0

        for index, row in notes_pd.iterrows():
            note_date = self.str_to_time(row['dateStr'])
            column = 'low_treatment' if 'low' in row['note'].lower() else 'more_aggressive'
            self.apply_adjustment_flags_to_row(aismb_df, note_date, column)

        aismb_df.to_csv(aismb_records_file_str, index=False)
    
    def apply_adjustment_flags_to_row(self, aismb_df, note_date, column):
        # search for multiple lines to find an existing matching timestamp
        row_to_update_index = self.find_row_index_for_date(aismb_df, note_date)
        if row_to_update_index is not None and row_to_update_index != -1:
            aismb_df.loc[aismb_df.iloc[row_to_update_index], column] = 1
        # for offset in range(10):
        #     note_date_plus_offset = note_date + timedelta(minutes=offset)
        #     note_date_str = self.time_to_str(note_date_plus_offset)
        #     if note_date_str in aismb_df['dateStr'].values:
        #         aismb_df.loc[aismb_df['dateStr'] == note_date_str, column] = 1
        #         return
    
    def find_row_index_for_date(self, df, date):
        for offset in range(10):
            date_plus_offset = date + timedelta(minutes=offset)
            date_str = self.time_to_str(date_plus_offset)
            if date_str in df['dateStr'].values:
                return df.index[df['dateStr'] == date_str].tolist()[0]

    
    def temp_add_tags(self):
        df_tags = pd.read_excel('androidaps_db-userEntry.xlsx','values_to_use')
        df_data = pd.read_excel('data.xlsx', 'training_data')

        for index, row in df_tags.iterrows():
            date = self.str_to_time(row['dateStr'])
            df_data_row_index = self.find_row_index_for_date(df_data, date)
            if df_data_row_index is not None:
                tags = row['Note']
                df_data.loc[df_data_row_index:df_data_row_index+11, 'tags0to60min'] = tags
                df_data.loc[df_data_row_index+12:df_data_row_index+23, 'tags60to120min'] = tags
                df_data.loc[df_data_row_index+24:df_data_row_index+35, 'tags120to180min'] = tags
                df_data.loc[df_data_row_index+36:df_data_row_index+47, 'tags180to240min'] = tags
            
            # for i in range(12):
            #     df_data.loc[df_data.iloc[df_data_row_index+i+12], 'tags60to120min'] = tags
            # for i in range(12):
            #     df_data.loc[df_data.iloc[df_data_row_index+i+24], 'tags120to180min'] = tags
            # for i in range(12):
            #     df_data.loc[df_data.iloc[df_data_row_index+i+36], 'tags180to240min'] = tags

        book = load_workbook('data.xlsx')
        writer = pd.ExcelWriter('data.xlsx', engine='openpyxl')
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}
        # start_row = writer.sheets['training_data'].max_row - 1
        # data = pd.read_excel('data.xlsx','training_data')
        # start_row = data.index[df['dateStr'] == start_date_time_str].tolist()[0]
        # 6/29/22 03:10PM - 9850
        df_data.to_excel(writer, sheet_name='training_data_2', index = False, header= False)

        writer.save()

        